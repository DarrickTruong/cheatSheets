# openpyxl excel
# in python cli (terminal)


import openpyxl
import os #change directory
from openpyxl.utils import get_column_letter
openpyxl.__version__ # to get version

os.chdir("your folder directory")

wb = openpyxl.load_workbook("file name here")
wb.sheetnames  #get sheet names
sheet = wb['Sheet name here']

type(wb) # find out what 'wb' is
wb.get_sheet_names() # get all sheets in wb

sheet.max_row/column # how many rows/column in sheet
get_column_letter(index) # get the column letter of index

# excel indexes start at 1 NOT 0 
    sheet.cell(row=i, column=52).value
    sheet['A2'].value